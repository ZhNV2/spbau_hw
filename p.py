import openpyxl
import sys
import os
import getpass
import email, smtplib, ssl

from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


table_file = sys.argv[1]
scand_dir = sys.argv[2]
email_file = sys.argv[3]


def email_format_error(text):
    print('Error: ' + text)
    print('Email file example:')
    print(email_file_template)
    sys.exit(0)


email_file_template = """
From: receiver@gmail.com
Subject: olympiad
Hello, dear patricipants!
lalaalal

Sincirely yours,
Boris
"""
table_format = """
id, шифр, фамилия, имя, отчество, e-mail
"""

def check_column_name(sheet, i, name):
    if sheet.cell(row=1, column=i).value.lower() != name:
        print('Error: ' + str(i) + '\'th column is not ' + name)
        exit(0)

def check_email_line(text, n, line):
    text = text + ':'
    qtext = '\"' + text + '\"'
    if len(line) == 0:
        email_format_error('missed ' + qtext + ' in '+ n +' line')
    if line[0] != text:
        email_format_error(n + ' line of a email file should start from ' + qtext)
    if len(line) == 1:
        email_format_error('missed text after ' + qtext)

def get_message_info(email_file):
    with open(email_file, 'r') as efile:
        sender_line = efile.readline().split()
        check_email_line('From', 'first', sender_line)
        subject_line = efile.readline().split()
        check_email_line('Subject', 'second', subject_line)
        body = efile.read()
    return sender_line[1], subject_line[1], body


sender, subject, body = get_message_info(email_file)
password = getpass.getpass(prompt='Password for ' + sender + ': ')

book = openpyxl.load_workbook(table_file)
sheet = book.active



ID_COLUMN = 1
CODE_COLUMN = 2
SURNAME_COLUMN = 3
NAME_COLUMN = 4
FAM_COULMN = 5
EMAIL_COLUMN = 6

check_column_name(sheet, ID_COLUMN, 'id')
check_column_name(sheet, CODE_COLUMN, 'шифр')
check_column_name(sheet, SURNAME_COLUMN, 'фамилия')
check_column_name(sheet, NAME_COLUMN, 'имя')
check_column_name(sheet, FAM_COULMN, 'отчество')
check_column_name(sheet, EMAIL_COLUMN, 'e-mail')

if scand_dir[-1] != '/':
        scand_dir = scand_dir + '/'

sent_column = 20
i = 2
SENT_MARK = 1
while sheet.cell(row=i, column = ID_COLUMN).value != '':
    surname = sheet.cell(row=i, column=SURNAME_COLUMN).value
    name = sheet.cell(row=i, column=NAME_COLUMN).value
    receiver = sheet.cell(row=i, column=EMAIL_COLUMN).value
    code = sheet.cell(row=i, column=CODE_COLUMN).value

    if sheet.cell(row=i, column=sent_column).value != SENT_MARK:
        filename = str(int(code)) + '.txt'
        scan_path = scand_dir + filename
        if not os.path.isfile(scan_path):
            print('Error: no scan found found for ' + surname + ' ' + name)
            exit(0)
        message = MIMEMultipart()
        message["From"] = sender
        message["To"] = receiver
        message["Subject"] = subject
        message["Bcc"] = receiver
        password = 'pomogi55'

        message.attach(MIMEText(body, "plain"))
        
        with open(scan_path, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition","attachment; filename=" + filename)
        message.attach(part)
        text = message.as_string()
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender, password)
            server.sendmail(sender, receiver, text)

        sheet.cell(row=i, column=sent_column).value = SENT_MARK
        book.save(table_file)
        print('sent to ' + surname + ' ' + name)
    else:
        print(surname + ' ' + name + ' skipped')
    i += 1
